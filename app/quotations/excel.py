"""Build a quotation-style Excel workbook from the Euroglass template."""

import os
import re
from datetime import date, datetime
from io import BytesIO

from openpyxl import load_workbook
from openpyxl.styles import Alignment

TEMPLATE_PATH = os.path.join(
    os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..")),
    "static",
    "excel",
    "QUOTATION.xlsx",
)

FIRST_LINE_ROW = 16
LAST_LINE_ROW = 38
MAX_LINE_ROWS = LAST_LINE_ROW - FIRST_LINE_ROW + 1


def calculated_sqft(width, height, qty):
    width = float(width or 0)
    height = float(height or 0)
    qty = float(qty or 0)
    if width > 0 and height > 0:
        return (width * height / 144.0) * qty
    return qty


def sqft_for_line(width, height, qty, sqft=None):
    if sqft not in (None, ""):
        custom = float(sqft)
        if custom > 0:
            return custom
    return calculated_sqft(width, height, qty)


def uses_custom_sqft(width, height, qty, sqft):
    try:
        entered = float(sqft or 0)
    except (TypeError, ValueError):
        return False
    if entered <= 0:
        return False
    return abs(entered - calculated_sqft(width, height, qty)) > 0.005


def line_amount(width, height, qty, rate, sqft=None):
    return sqft_for_line(width, height, qty, sqft) * float(rate or 0)


def _as_date(value):
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    value = str(value or "").strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%Y-%m-%d %H:%M:%S"):
        try:
            return datetime.strptime(value[:19], fmt).date()
        except ValueError:
            continue
    return datetime.now().date()


def _set_number(cell, value):
    if value is None or value == "":
        cell.value = None
        return
    number = float(value)
    if number == 0:
        cell.value = None
        return
    cell.value = int(number) if number == int(number) else number


def quotation_download_name(customer_name, extension="xlsx"):
    """Safe download filename from the quotation Name field."""
    base = str(customer_name or "").strip()
    base = re.sub(r'[<>:"/\\|?*\x00-\x1f]+', "", base)
    base = re.sub(r"\s+", "_", base).strip("._")
    stem = base or "QUOTATION"
    if len(stem) > 80:
        stem = stem[:80].rstrip("_")
    return f"{stem}.{extension.lstrip('.')}"


def _set_description(cell, value, sheet, row_num):
    text = str(value or "").strip()
    cell.value = text or None
    cell.alignment = Alignment(wrap_text=True, vertical="top")
    if not text:
        sheet.row_dimensions[row_num].height = 15
        return
    line_breaks = text.count("\n") + 1
    wrapped_lines = max(line_breaks, (len(text) // 42) + 1)
    sheet.row_dimensions[row_num].height = min(120, max(18, 15 * wrapped_lines))


def build_quotation_xlsx(header, lines):
    """Return a BytesIO workbook named in quotation style."""
    if not os.path.exists(TEMPLATE_PATH):
        raise FileNotFoundError(f"Quotation template not found: {TEMPLATE_PATH}")

    workbook = load_workbook(TEMPLATE_PATH)
    if "LEDGER" in workbook.sheetnames:
        del workbook["LEDGER"]

    sheet = workbook["INVOICE"] if "INVOICE" in workbook.sheetnames else workbook.active
    sheet.title = "QUOTATION"
    if sheet.column_dimensions["B"].width is None or sheet.column_dimensions["B"].width < 28:
        sheet.column_dimensions["B"].width = 42

    sheet["C4"] = header.get("quotation_no")
    sheet["C5"] = _as_date(header.get("quotation_date"))
    sheet["C6"] = header.get("customer_name") or ""
    sheet["C7"] = header.get("address") or ""
    sheet["C8"] = header.get("project") or ""
    sheet["C9"] = header.get("work_type") or ""
    sheet["C10"] = header.get("engineer") or ""
    sheet["C11"] = header.get("contact_no") or ""

    note_lines = [part.strip() for part in str(header.get("notes") or "").splitlines() if part.strip()]
    sheet["B41"] = note_lines[0] if len(note_lines) > 0 else None
    sheet["B42"] = note_lines[1] if len(note_lines) > 1 else None
    sheet["B43"] = " ".join(note_lines[2:]) if len(note_lines) > 2 else None

    used_lines = list(lines)[:MAX_LINE_ROWS]
    for offset in range(MAX_LINE_ROWS):
        row = FIRST_LINE_ROW + offset
        if offset < len(used_lines):
            line = used_lines[offset]
            width = line.get("width") or 0
            height = line.get("height") or 0
            qty = line.get("quantity") or line.get("qty") or 0
            rate = line.get("rate") or 0
            sqft = line.get("sqft")

            sheet[f"A{row}"] = offset + 1
            _set_description(sheet[f"B{row}"], line.get("description") or line.get("item_name") or "", sheet, row)
            _set_number(sheet[f"I{row}"], width)
            _set_number(sheet[f"J{row}"], height)
            _set_number(sheet[f"K{row}"], qty)
            _set_number(sheet[f"M{row}"], rate)

            if uses_custom_sqft(width, height, qty, sqft):
                sheet[f"L{row}"] = round(float(sqft), 4)
            elif float(width or 0) > 0 and float(height or 0) > 0:
                sheet[f"L{row}"] = f"=((I{row}*J{row})/144)*K{row}"
            else:
                sheet[f"L{row}"] = f"=IF(K{row}=\"\",\"\",K{row})"
            sheet[f"N{row}"] = f"=IF(L{row}=\"\",\"\",L{row}*M{row})"
        else:
            sheet[f"A{row}"] = None
            sheet[f"B{row}"] = None
            sheet[f"I{row}"] = None
            sheet[f"J{row}"] = None
            sheet[f"K{row}"] = None
            sheet[f"M{row}"] = None
            sheet[f"L{row}"] = f"=((I{row}*J{row})/144)*K{row}"
            sheet[f"N{row}"] = f"=L{row}*M{row}"

    advance = float(header.get("advance") or 0)
    sheet["N42"] = advance

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output
