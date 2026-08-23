import io
import re
from datetime import date

from openpyxl import Workbook, load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from app.validators import ValidationErrors, clean_string, clean_optional_string, clean_positive_decimal, clean_positive_int

HEADER_ALIASES = {
    "item_name": {"item_name", "itemname", "item", "name", "product", "product_name"},
    "brand": {"brand", "brand_name", "brandname", "make", "manufacturer"},
    "category": {"category", "category_name", "categoryname"},
    "supplier_name": {"supplier", "supplier_name", "suppliername"},
    "purchase_rate": {"purchase_rate", "purchaserate", "purchase", "purchase_price", "cost"},
    "sale_rate": {"sale_rate", "salerate", "sale", "sale_price", "price"},
    "qty": {"qty", "quantity", "stock", "amount"},
}

REQUIRED_FIELDS = {"item_name", "category", "supplier_name", "purchase_rate", "sale_rate", "qty"}
OPTIONAL_FIELDS = {"brand"}


def _normalize_header(value):
    if value is None:
        return ""
    return re.sub(r"[\s\-]+", "_", str(value).strip().lower())


def _detect_columns(header_row):
    mapping = {}
    for index, cell in enumerate(header_row):
        key = _normalize_header(cell)
        for field, aliases in HEADER_ALIASES.items():
            if key in aliases and field not in mapping:
                mapping[field] = index
    return mapping


def _validate_row(row_number, row_data, errors):
    return {
        "item_name": clean_string(
            row_data.get("item_name"), f"row_{row_number}_item_name", errors, max_len=100, label=f"Row {row_number} item name"
        ),
        "brand": clean_optional_string(
            row_data.get("brand"), f"row_{row_number}_brand", errors, max_len=100, label=f"Row {row_number} brand"
        ),
        "category": clean_string(
            row_data.get("category"), f"row_{row_number}_category", errors, max_len=50, label=f"Row {row_number} category"
        ),
        "supplier_name": clean_string(
            row_data.get("supplier_name"), f"row_{row_number}_supplier_name", errors, max_len=60, label=f"Row {row_number} supplier name"
        ),
        "purchase_rate": clean_positive_decimal(
            row_data.get("purchase_rate"), f"row_{row_number}_purchase_rate", errors, label=f"Row {row_number} purchase rate"
        ),
        "sale_rate": clean_positive_decimal(
            row_data.get("sale_rate"), f"row_{row_number}_sale_rate", errors, label=f"Row {row_number} sale rate"
        ),
        "qty": clean_positive_int(
            row_data.get("qty"), f"row_{row_number}_qty", errors, min_val=0, label=f"Row {row_number} quantity"
        ),
    }


def _cell_value(value):
    if value is None:
        return ""
    return str(value).strip()


def parse_items_xlsx(file_stream):
    workbook = None

    try:
        if hasattr(file_stream, "seek"):
            file_stream.seek(0)

        workbook = load_workbook(file_stream, read_only=True, data_only=True)
        sheet = workbook.active
        rows_iter = sheet.iter_rows(values_only=True)

        first_non_empty = None
        for row in rows_iter:
            if any(cell is not None and str(cell).strip() for cell in row):
                first_non_empty = row
                break

        if first_non_empty is None:
            raise ValueError("The Excel file is empty.")

        required_fields = REQUIRED_FIELDS
        header_map = _detect_columns(first_non_empty)

        if header_map and required_fields.issubset(header_map):
            start_row_number = 2
            data_rows_iter = rows_iter
        else:
            if len(first_non_empty) >= 7:
                header_map = {
                    "item_name": 0,
                    "brand": 1,
                    "category": 2,
                    "supplier_name": 3,
                    "purchase_rate": 4,
                    "sale_rate": 5,
                    "qty": 6,
                }
                start_row_number = 1

                def _prepend_first_row():
                    yield first_non_empty
                    for next_row in rows_iter:
                        yield next_row

                data_rows_iter = _prepend_first_row()
            elif len(first_non_empty) >= 6:
                header_map = {
                    "item_name": 0,
                    "category": 1,
                    "supplier_name": 2,
                    "purchase_rate": 3,
                    "sale_rate": 4,
                    "qty": 5,
                }
                start_row_number = 1

                def _prepend_first_row():
                    yield first_non_empty
                    for next_row in rows_iter:
                        yield next_row

                data_rows_iter = _prepend_first_row()
            else:
                raise ValueError(
                    "Could not detect column headers. Use columns: ItemName, Brand, Category, "
                    "SupplierName, PurchaseRate, SaleRate, Qty."
                )

        parsed_fields = set(required_fields)
        parsed_fields.update(field for field in OPTIONAL_FIELDS if field in header_map)

        valid_items = []
        row_errors = []

        for offset, row in enumerate(data_rows_iter):
            row_index = start_row_number + offset
            row_data = {
                field: _cell_value(row[header_map[field]] if header_map[field] < len(row) else "")
                for field in parsed_fields
            }

            if not any(row_data.values()):
                continue

            errors = ValidationErrors()
            validated = _validate_row(row_index, row_data, errors)

            if errors.valid:
                valid_items.append(validated)
            else:
                row_errors.append(f"Row {row_index}: {errors.first()}")

        if not valid_items and row_errors:
            raise ValueError(row_errors[0])

        if not valid_items:
            raise ValueError("No valid item rows were found in the Excel file.")

        return valid_items, row_errors

    except InvalidFileException as exc:
        raise ValueError("The uploaded file is not a valid .xlsx workbook.") from exc
    finally:
        if workbook is not None:
            workbook.close()


def import_items(app, items):
    db = None
    cursor = None
    inserted = 0
    updated = 0
    row_errors = []

    try:
        from app.db import get_db_connection
        from app.items.routes import _ensure_item_brand_column
        from app.tenancy import next_owner_no, next_table_id, owner_sql, request_user_id

        db = get_db_connection(app)
        cursor = db.cursor()
        _ensure_item_brand_column(db, cursor)

        for item in items:
            cursor.execute(
                f"""
                SELECT TOP 1 CategoryID, CategoryName
                FROM Category
                WHERE LOWER(LTRIM(RTRIM(CategoryName))) = LOWER(LTRIM(RTRIM(?)))
                  AND {owner_sql()}
                """,
                (item["category"],),
            )
            category = cursor.fetchone()

            if not category:
                row_errors.append(f"Category not found: {item['category']}")
                continue

            cursor.execute(
                f"""
                SELECT TOP 1 SupplierID, SupplierName
                FROM Supplier
                WHERE LOWER(LTRIM(RTRIM(SupplierName))) = LOWER(LTRIM(RTRIM(?)))
                  AND {owner_sql()}
                """,
                (item["supplier_name"],),
            )
            supplier = cursor.fetchone()

            if not supplier:
                row_errors.append(f"Supplier not found: {item['supplier_name']}")
                continue

            cursor.execute(
                f"""
                SELECT TOP 1 ItemID
                FROM Item
                WHERE LOWER(LTRIM(RTRIM(ItemName))) = LOWER(LTRIM(RTRIM(?)))
                  AND CategoryID = ?
                  AND {owner_sql()}
                ORDER BY Qty DESC, ItemID ASC
                """,
                (item["item_name"], category.CategoryID),
            )
            existing_item = cursor.fetchone()

            if existing_item:
                item_id = existing_item.ItemID
                cursor.execute(
                    f"""
                    UPDATE Item
                    SET Qty = Qty + ?, PurchaseRate = ?, SaleRate = ?,
                        Brand = COALESCE(?, Brand)
                    WHERE ItemID = ? AND {owner_sql()}
                    """,
                    (
                        item["qty"],
                        item["purchase_rate"],
                        item["sale_rate"],
                        item.get("brand"),
                        existing_item.ItemID,
                    ),
                )
                updated += 1
            else:
                next_item_id = next_table_id(cursor, "Item", "ItemID")
                next_item_no = next_owner_no(cursor, "Item", "ItemNo")
                cursor.execute(
                    """
                    INSERT INTO Item (ItemID, ItemNo, ItemName, Brand, CategoryID, PurchaseRate, SaleRate, Qty, UserID)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        next_item_id,
                        next_item_no,
                        item["item_name"],
                        item.get("brand"),
                        category.CategoryID,
                        item["purchase_rate"],
                        item["sale_rate"],
                        item["qty"],
                        request_user_id(),
                    ),
                )
                item_id = next_item_id
                inserted += 1

            total = item["qty"] * item["purchase_rate"]

            # PurchaseID is assigned explicitly everywhere else in the app
            # (see purchases/routes.py), which leaves Postgres's own SERIAL
            # sequence lagging behind. Relying on that sequence's default
            # here caused "duplicate key" errors once enough manually
            # created purchases outpaced it, so assign the id the same way.
            next_purchase_id = next_table_id(cursor, "Purchases", "PurchaseID")
            cursor.execute(
                """
                INSERT INTO Purchases (PurchaseID, PurchaseDate, SupplierID, TotalAmount, UserID)
                OUTPUT INSERTED.PurchaseID
                VALUES (?, ?, ?, ?, ?)
                """,
                (next_purchase_id, date.today(), supplier.SupplierID, total, request_user_id()),
            )
            purchase_id = int(cursor.fetchone()[0])

            cursor.execute(
                """
                INSERT INTO PurchaseDetails (PurchaseID, ItemID, Particulars, Qty, PurchaseRate)
                VALUES (?, ?, ?, ?, ?)
                """,
                (purchase_id, item_id, item["item_name"], item["qty"], item["purchase_rate"]),
            )

        db.commit()
        return inserted, updated, row_errors

    except Exception:
        if db is not None:
            db.rollback()
        raise

    finally:
        if cursor is not None:
            cursor.close()


def build_items_template_xlsx():
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Items"
    sheet.append(["ItemName", "Brand", "Category", "SupplierName", "PurchaseRate", "SaleRate", "Qty"])
    sheet.append(["Hammer", "Stanley", "Glass Hardware 12MM", "ABC Supplier", 150.00, 200.00, 25])
    sheet.append(["Screwdriver Set", "Bosch", "Aluminium Hardware", "ABC Supplier", 80.50, 120.00, 40])

    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


BRAND_UPDATE_ALIASES = {
    "item_name": {"item_name", "itemname", "item", "name", "product", "product_name"},
    "brand": {"brand", "brand_name", "brandname", "make", "manufacturer"},
    "category": {"category", "category_name", "categoryname"},
}


def _detect_brand_columns(header_row):
    mapping = {}
    for index, cell in enumerate(header_row):
        key = _normalize_header(cell)
        for field, aliases in BRAND_UPDATE_ALIASES.items():
            if key in aliases and field not in mapping:
                mapping[field] = index
    return mapping


def parse_brands_xlsx(file_stream):
    """Parse a two-column (or three with Category) sheet for brand-only bulk update."""
    workbook = None
    try:
        if hasattr(file_stream, "seek"):
            file_stream.seek(0)
        workbook = load_workbook(file_stream, read_only=True, data_only=True)
        sheet = workbook.active
        rows_iter = sheet.iter_rows(values_only=True)

        first_non_empty = None
        for row in rows_iter:
            if any(cell is not None and str(cell).strip() for cell in row):
                first_non_empty = row
                break
        if first_non_empty is None:
            raise ValueError("The Excel file is empty.")

        header_map = _detect_brand_columns(first_non_empty)
        if {"item_name", "brand"}.issubset(header_map):
            start_row_number = 2
            data_rows_iter = rows_iter
        elif len(first_non_empty) >= 2:
            header_map = {"item_name": 0, "brand": 1}
            if len(first_non_empty) >= 3:
                header_map["category"] = 2
            start_row_number = 1

            def _prepend_first_row():
                yield first_non_empty
                for next_row in rows_iter:
                    yield next_row

            data_rows_iter = _prepend_first_row()
        else:
            raise ValueError("Use columns: ItemName, Brand (Category optional).")

        valid_rows = []
        row_errors = []
        for offset, row in enumerate(data_rows_iter, start=start_row_number):
            if not row or not any(cell is not None and str(cell).strip() for cell in row):
                continue
            row_data = {
                field: _cell_value(row[index]) if index < len(row) else ""
                for field, index in header_map.items()
            }
            errors = ValidationErrors()
            item_name = clean_string(
                row_data.get("item_name"),
                f"row_{offset}_item_name",
                errors,
                max_len=100,
                label=f"Row {offset} item name",
            )
            brand = clean_optional_string(
                row_data.get("brand"),
                f"row_{offset}_brand",
                errors,
                max_len=100,
                label=f"Row {offset} brand",
            )
            category = clean_optional_string(
                row_data.get("category"),
                f"row_{offset}_category",
                errors,
                max_len=50,
                label=f"Row {offset} category",
            )
            if errors.valid:
                if not brand:
                    row_errors.append(f"Row {offset}: brand is empty — skipped.")
                    continue
                valid_rows.append({"item_name": item_name, "brand": brand, "category": category or None})
            else:
                row_errors.append(errors.first())

        if not valid_rows and not row_errors:
            raise ValueError("No data rows found in the Excel file.")
        return valid_rows, row_errors
    finally:
        if workbook is not None:
            workbook.close()


def update_item_brands(app, rows):
    """Update Brand only — no qty, rates, or purchase records touched."""
    from app.db import get_db_connection
    from app.items.routes import _ensure_item_brand_column
    from app.tenancy import owner_sql

    db = get_db_connection(app)
    cursor = db.cursor()
    updated = 0
    skipped = 0
    row_errors = []
    try:
        _ensure_item_brand_column(db, cursor)
        for row in rows:
            params = [row["item_name"]]
            category_sql = ""
            if row.get("category"):
                category_sql = """
                  AND CategoryID IN (
                      SELECT CategoryID FROM Category
                      WHERE LOWER(LTRIM(RTRIM(CategoryName))) = LOWER(LTRIM(RTRIM(?)))
                        AND {owner}
                  )
                """.format(owner=owner_sql())
                params.append(row["category"])

            cursor.execute(
                f"""
                SELECT ItemID, ItemName
                FROM Item
                WHERE LOWER(LTRIM(RTRIM(ItemName))) = LOWER(LTRIM(RTRIM(?)))
                  AND {owner_sql()}
                  {category_sql}
                ORDER BY ItemID ASC
                """,
                tuple(params),
            )
            matches = cursor.fetchall()
            if not matches:
                row_errors.append(f"Item not found: {row['item_name']}")
                continue
            if len(matches) > 1 and not row.get("category"):
                row_errors.append(
                    f"Multiple items named \"{row['item_name']}\" — add Category column to pick the right one."
                )
                skipped += 1
                continue

            for match in matches:
                cursor.execute(
                    f"UPDATE Item SET Brand = ? WHERE ItemID = ? AND {owner_sql()}",
                    (row["brand"], match.ItemID),
                )
                updated += int(cursor.rowcount or 0)
        db.commit()
        return updated, skipped, row_errors
    except Exception:
        db.rollback()
        raise
    finally:
        cursor.close()


def build_brands_template_xlsx():
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Brands"
    sheet.append(["ItemName", "Brand", "Category"])
    sheet.append(["Euro 38 Floor hinge machine", "Dorma", "Glass Hardware 12MM"])
    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


def build_brands_export_xlsx(app):
    """Export all items with current names/categories so user can fill Brand column."""
    from app.db import execute_query
    from app.tenancy import owner_sql

    rows = execute_query(
        app,
        f"""
        SELECT i.ItemName, COALESCE(i.Brand, '') AS Brand, COALESCE(c.CategoryName, '') AS CategoryName
        FROM Item i
        LEFT JOIN Category c ON c.CategoryID = i.CategoryID
        WHERE {owner_sql("i")}
        ORDER BY COALESCE(i.ItemNo, i.ItemID), i.ItemID
        """,
    )
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Brands"
    sheet.append(["ItemName", "Brand", "Category"])
    for row in rows:
        sheet.append([row.ItemName, row.Brand or "", row.CategoryName or ""])
    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    return output
